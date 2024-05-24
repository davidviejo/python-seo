import requests
from bs4 import BeautifulSoup
from collections import Counter
import openpyxl as xl

def obtener_serps(palabra_clave):
    url = f"https://www.google.com/search?q={palabra_clave}&num=5"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")
    serps = []
    for result in soup.find_all("div", class_="yuRUbf"):
        link = result.find("a")
        serps.append(link["href"])
    return serps

def obtener_contenido(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    texto = soup.get_text(separator=" ")
    return texto

def contar_palabras(texto):
    palabras = texto.split()
    contador = Counter(palabras)
    return contador

def contar_pares_palabras(texto):
    palabras = texto.split()
    pares = [f"{palabras[i]} {palabras[i+1]}" for i in range(len(palabras)-1)]
    contador = Counter(pares)
    return contador

def contar_imagenes(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    imagenes = soup.find_all("img")
    return len(imagenes)

def extraer_encabezados(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    encabezados = []
    for etiqueta in ["h1", "h2", "h3"]:
        elementos = soup.find_all(etiqueta)
        for elemento in elementos:
            encabezado = f"{etiqueta.upper()}: {elemento.text.strip()}"
            encabezados.append(encabezado)
    return encabezados

def extraer_atributo_alt(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    imagenes = soup.find_all("img")
    atributos_alt = []
    for imagen in imagenes:
        src = imagen.get("src", "")
        alt = imagen.get("alt", "")
        atributos_alt.append((src, alt))
    return atributos_alt

# Solicitar la palabra clave al usuario
palabra_clave = input("Ingrese la palabra clave: ")

# Obtener las SERPs
serps = obtener_serps(palabra_clave)

# Crear un archivo Excel
libro = xl.Workbook()
hoja_total = libro.active
hoja_total.title = "Conteo Palabras"

# Crear una hoja para encabezados
hoja_encabezados = libro.create_sheet(title="Encabezados")

# Título de las columnas para la hoja de encabezados
hoja_encabezados.cell(row=1, column=1).value = "URL"
hoja_encabezados.cell(row=1, column=2).value = "Encabezado"
hoja_encabezados.cell(row=1, column=3).value = "Tipo de encabezado"

# La variable para controlar la fila actual en la hoja de encabezados
fila_encabezados = 2

# Crear una hoja para los atributos alt de las imágenes
hoja_atributos_alt = libro.create_sheet(title="Atributos alt de imágenes")

# Título de las columnas para la hoja de atributos alt
hoja_atributos_alt.cell(row=1, column=1).value = "Competidor"
hoja_atributos_alt.cell(row=1, column=2).value = "URL de la imagen"
hoja_atributos_alt.cell(row=1, column=3).value = "Alt"

# La variable para controlar la fila actual en la hoja de atributos alt
fila_atributos_alt = 2

# Recorrer cada SERP
for i, serp_url in enumerate(serps, start=1):
    # Obtener el contenido de la URL
    contenido = obtener_contenido(serp_url)
    
    # Contar palabras y pares de palabras
    contador_palabras = contar_palabras(contenido)
    contador_pares = contar_pares_palabras(contenido)
    
    # Extraer encabezados
    encabezados = extraer_encabezados(serp_url)
    
    # Guardar los encabezados en la hoja de encabezados
    for encabezado in encabezados:
        hoja_encabezados.cell(row=fila_encabezados, column=1).value = serp_url
        hoja_encabezados.cell(row=fila_encabezados, column=2).value = encabezado.split(": ")[1]
        hoja_encabezados.cell(row=fila_encabezados, column=3).value = encabezado.split(": ")[0]
        fila_encabezados += 1

    # Guardar los resultados en la hoja correspondiente
    hoja = libro.create_sheet(title=f"SERP {i}")
    hoja.cell(row=1, column=1).value = "Palabra"
    hoja.cell(row=1, column=2).value = "Conteo"
    hoja.cell(row=1, column=3).value = "Par de palabras"
    hoja.cell(row=1, column=4).value = "Conteo"
    
    row = 2
    for palabra, conteo in contador_palabras.items():
        hoja.cell(row=row, column=1).value = palabra
        hoja.cell(row=row, column=2).value = conteo
        row += 1
    
    row = 2
    for par, conteo in contador_pares.items():
        hoja.cell(row=row, column=3).value = par
        hoja.cell(row=row, column=4).value = conteo
        row += 1
    
    # Contar imágenes
    num_imagenes = contar_imagenes(serp_url)
    hoja.cell(row=1, column=5).value = "Número de imágenes"
    hoja.cell(row=2, column=5).value = num_imagenes

    # Obtener los atributos alt de las imágenes
    atributos_alt = extraer_atributo_alt(serp_url)
    
    # Guardar los atributos alt en la hoja de atributos alt
    for src, alt in atributos_alt:
        hoja_atributos_alt.cell(row=fila_atributos_alt, column=1).value = serp_url
        hoja_atributos_alt.cell(row=fila_atributos_alt, column=2).value = src
        hoja_atributos_alt.cell(row=fila_atributos_alt, column=3).value = alt
        fila_atributos_alt += 1

    # Agrupar la información de conteo de palabras
    hoja_total.cell(row=2, column=1).value = "URL"
    hoja_total.cell(row=2, column=i+1).value = serp_url
    hoja_total.cell(row=3, column=1).value = "Palabra"
    hoja_total.cell(row=3, column=i+1).value = "Conteo"
    hoja_total.cell(row=4, column=1).value = "Par de palabras"
    hoja_total.cell(row=4, column=i+1).value = "Conteo"
    
    row = 5
    for par, conteo in contador_pares.items():
        hoja_total.cell(row=row, column=1).value = par
        hoja_total.cell(row=row, column=i+1).value = conteo
        row += 1

# Guardar el archivo Excel con el nombre "palabra clave buscada + analisis"
nombre_archivo = f"{palabra_clave} - analisis.xlsx"
libro.save(nombre_archivo)
